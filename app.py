import pandas as pd
from neo4j import GraphDatabase
import numpy as np
import datetime
import xlsxwriter
import os
from openpyxl import load_workbook

src_uri = "neo4j+s://174cd36c.databases.neo4j.io:7687"
src_user = "neo4j"
src_password = "w21V4bw-6kTp9hceHMbnlt5L9X1M4upuuq2nD7tD_xU"
uri = "bolt://localhost:7687"
user = "neo4j"
password = "2310"


# src_uri = "neo4j://20.107.79.39:7687"
# src_user = "neo4j"
# src_password = "Accelerati0n"


class App:
    def __init__(self, uri_, user_, password_):
        self.driver = GraphDatabase.driver(uri_, auth=(user_, password_))

    def close(self):
        # Don't forget to close the driver connection when you are finished with it
        self.driver.close()

    def load_data(self, remote_uri, remote_user, remote_pswd):
        q_data_obtain = '''
            MATCH (n)-[r]->(m)
            RETURN n.name AS n_name, n.id AS n_id, properties(r).weight AS weight, m.name AS m_name, m.id AS m_id
            '''
        q_create = '''
            LOAD CSV WITH HEADERS FROM 'file:///2.csv' AS row
            MERGE (n:Work {id: row.n_id, name: row.n_name})
            MERGE (m:Work {id: row.m_id, name: row.m_name})
            CREATE (n)-[r:FOLLOWS {weight: row.weight}]->(m);
            '''
        # obtaining data
        src_driver = GraphDatabase.driver(remote_uri, auth=(remote_user, remote_pswd))
        print('open')
        result = src_driver.session().run(q_data_obtain).data()
        src_driver.close()
        df = pd.DataFrame(result)
        save_path = 'C:\\Users\\Nikita\\.Neo4jDesktop\\relate-data\\dbmss\\dbms-44d5c24b-bb41-4e4b-bbeb-f18bca851f09' \
                    '\\import\\'
        df.to_csv(save_path + '2.csv', index=False)
        with self.driver.session() as session:
            session.run("MATCH (n) DETACH DELETE n")  # Очистка
            session.run(q_create)

    def removing_node(self, id: str):
        income_data_obtain = '''
            MATCH (n)-[]->(m)
            WHERE m.id = $id
            RETURN n
            '''
        outcome_data_obtain = '''
            MATCH (n)-[]->(m)
            WHERE n.id = $id
            RETURN m
            '''
        with self.driver.session() as session:
            incoming = session.run(income_data_obtain, id=id).data()
            outcoming = session.run(outcome_data_obtain, id=id).data()
            # преобразование результатов запроса в numpy.array
            incoming = np.array([row['n']['id'] for row in incoming])
            outcoming = np.array([row['m']['id'] for row in outcoming])

            for element in incoming:
                for subelement in outcoming:
                    session.run('''
                                        MERGE (n:Work {id: $id1})
                                        MERGE (m:Work {id: $id2})
                                        MERGE (n)-[r:FOLLOWS]->(m)
                                        ''',
                                id1=element,
                                id2=subelement
                                )
            session.run("MATCH (n) WHERE n.id = $id DETACH DELETE n", id=id)

    def get_all_id(self):
        q_data_obtain = '''
            MATCH (n)
            RETURN n
            '''
        result = self.driver.session().run(q_data_obtain).data()
        id_lst = []
        for i in result:
            id_lst.append((i['n']['id']))
        return list(set(id_lst))

    def new_graph(self, target_ids: list):
        for element in self.get_all_id():
            if element not in target_ids:
                self.removing_node(element)

    def result_to_excel(self):
        q_data_obtain = '''
                        MATCH (n)-[]->(m)
                        RETURN n, m
                        '''
        result = self.driver.session().run(q_data_obtain).data()
        n_id_arr = np.array([r['n']['id'] for r in result])
        m_id_arr = np.array([r['m']['id'] for r in result])
        n_name_arr = np.array([r['n']['name'] for r in result])
        m_name_arr = np.array([r['m']['name'] for r in result])
        preds = dict()
        flws = dict()
        names = dict()
        for n_id, n_name, m_id, m_name in zip(n_id_arr, n_name_arr, m_id_arr, m_name_arr):
            names[n_id] = n_name
            names[m_id] = m_name
            try:
                flws[n_id].add(m_id)
            except KeyError:
                flws[n_id] = {m_id}
            try:
                preds[m_id].add(n_id)
            except KeyError:
                preds[m_id] = {n_id}
        workbook = xlsxwriter.Workbook('result.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'Идентификатор операции')
        worksheet.set_column(0, 0, 25)
        worksheet.write(0, 1, 'Название операции')
        worksheet.set_column(1, 1, 75)
        worksheet.write(0, 2, 'Предшественники')
        worksheet.write(0, 3, 'Последователи')
        worksheet.set_column(2, 3, 30)
        row = 1
        for i in self.get_all_id():
            worksheet.write(row, 0, i)
            worksheet.write(row, 1, names[i])
            try:
                worksheet.write(row, 2, ', '.join(preds[i]))
            except KeyError:
                worksheet.write(row, 2, '')
            try:
                worksheet.write(row, 3, ', '.join(flws[i]))
            except KeyError:
                worksheet.write(row, 3, '')
            row += 1
        workbook.close()

    def del_extra_rel(self):
        q_delete = '''
            match (b)<-[r]-(a)-->(c)-->(b)
            delete r
            '''
        self.driver.session().run(q_delete)

    def ordering(self, filename):
        q_data_obtain = '''
        MATCH (n)-[]->(m)
        RETURN n.id AS n_id, m.id AS m_id
        '''
        file_name = 'data/solution_schedule.xlsx'
        schedDF = pd.read_excel(file_name, sheet_name="Состав работ", dtype=str, index_col=0)
        schedDF = schedDF[['wbs2', 'vendor_code']]  # , 'name']
        schedDF['precursors'] = ''
        schedDF['followers'] = ''
        wbs2_arr = schedDF.wbs2.unique()

        # checking
        # vendors = schedDF[schedDF.wbs2 == wbs2_arr[0]].vendor_code.to_numpy()
        # self.load_data(src_uri, src_user, src_password)
        # self.new_graph(vendors)
        # self.del_extra_rel()

        for wbs2 in reversed(wbs2_arr):
            wbs_df = schedDF[schedDF.wbs2 == wbs2]
            vendors = wbs_df.vendor_code.to_numpy()
            self.load_data(src_uri, src_user, src_password)
            self.new_graph(vendors)
            self.del_extra_rel()

            # Должно быть цикле
            result = self.driver.session().run(q_data_obtain).data()
            df = pd.DataFrame(result)
            for ind, vend in enumerate(vendors):
                ind2 = wbs_df.index[wbs_df.vendor_code == vend].tolist()[0]
                # if ind != ind2[0]:
                #     print(ind, ind2, vend)
                flwDF = df.loc[df.n_id == vend]
                if not flwDF.empty:
                    flws = flwDF.m_id.to_numpy()
                    wbs_df.at[ind2, 'followers'] = ', '.join(flws)

                predDF = df.loc[df.m_id == vend]
                if not predDF.empty:
                    preds = predDF.n_id.to_numpy()
                    wbs_df.at[ind2, 'precursors'] = ', '.join(preds)

            # book = load_workbook('data/new_sched.xlsx')
            with pd.ExcelWriter('new_sched.xlsx', engine='openpyxl', mode='a') as writer:
                # writer.book = book
                wbs_df.to_excel(writer, sheet_name=wbs2)
                writer.save()



def main():
    starttime = datetime.datetime.now()
    app = App(uri, user, password)
    # app.load_data(src_uri, src_user, src_password)
    app.ordering('data/solution_schedule.xlsx')
    print('data loaded', datetime.datetime.now() - starttime)
    # app.new_graph(['100203',  # Монтаж планка натягивающая
    #                '171670',  # Монтаж трос стен вi-l дл.5986 серьга вверху
    #                '192057',  # Монтаж алюминиевая направляющая для пола bi-level ii-уровень
    #                '111281',  # Монтаж балка несущая вi-l5м дл.
    #                '161564',  # Монтаж панель стеклянная
    #                '165160'  # Монтаж дверь двухстворчатая e6/ev1 din 17611
    #                ])
    # print('new_graph', datetime.datetime.now() - starttime)
    # app.del_extra_rel()
    # app.result_to_excel()
    # app.close()
    # print(datetime.datetime.now() - starttime)


if __name__ == "__main__":
    main()
